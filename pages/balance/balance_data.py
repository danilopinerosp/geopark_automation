from dotenv import load_dotenv
import numpy as np

from utils.functions import filter_data_by_date, load_companies, load_data, load_oil_types_names

from openpyxl import Workbook, load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.drawing.image import Image
from openpyxl.utils import get_column_letter
import pandas as pd

from utils.constants import months, conditions, operations
from dash import callback_context, html

import os

def get_date_last_report(data):
    try:
        return data['fecha'].max().strftime('%d/%m/%Y')
    except:
        return "Aún no hay datos"

def get_cumulated(data, start_date, end_date, operation_type, operation_condition, company):
    filtered_data = filter_data_by_date(data, start_date, end_date)
    if filtered_data.shape != (0,0):
        filtered_data = filtered_data[['empresa', 'operacion', operation_condition]]
        filter = (filtered_data['operacion'] == operation_type) & (filtered_data['empresa'] == company)
        cumulated = filtered_data[filter][operation_condition].sum()
    else:
        cumulated = 0
    return f"{cumulated:,.2f}"

def update_indicators(data, operation_type, operation_conditions):
    try:
        # Agrupar los valores del DataFrame y hacer una suma por cada grupo
        datos_agrupados = data.groupby(['fecha', 'empresa', 'operacion'])[operation_conditions].sum()
        # Seleccionar el GOV para el último día reportado
        last = np.round(datos_agrupados.unstack().unstack()[operation_type]['GEOPARK'][-1], 2)
        # Seleccionar el GOV para el penúltimo día reportado
        previous = np.round(datos_agrupados.unstack().unstack()[operation_type]['GEOPARK'][-2], 2)
    except:
        last = 0
        previous = 0
    return (last, previous)

def read_data_daily_reports(book, filename, start_cell=1, end_cell=270):
    """
    Leer los datos del documento que recibe como parámetro y retorna una
    lista de diccionarios con todos los datos de documento

    Parámetros:
    -----------
    documento -> str - Cadena de caracteres que contiene la ruta del
                        documento de donde se quieren leer los leer_datos
    Return:
    ------
    list -> Retorna una lista de diccionarios con los datos del documento.
    """
    companies = load_companies()
    oil_types_names = load_oil_types_names()
    sheet =  book.active #Por defecto toma como activa la primera hoja

    #Extraer le fecha del reporte del nombre del documento
    file_date = get_date_report(filename)

    list_data = list() # Para almacenar la lista de diccionarios

    # Recorrer todas las filas del documento excel y extraer los valores
    for i in range(start_cell, end_cell):
        column_b = 'B' + str(i) # ayuda a recorrer la columna B
        value_cell = sheet[column_b].value
        datos = dict()     # Almacena los datos por cada entrada
        if value_cell in companies:
            company = value_cell
            continue
        if value_cell in operations:
            operation = value_cell
            continue
        if value_cell in oil_types_names:
            # si valor se encuentra en la constante CAMPO guarda todos los datos
            # en el diccionario datos
            datos['fecha'] = file_date
            datos['empresa'] = company
            datos['operacion'] = operation
            datos['tipo crudo'] = value_cell
            datos['GOV'] = sheet['D' + str(i)].value
            datos['GSV'] = sheet['E' + str(i)].value
            datos['NSV'] = sheet['F' + str(i)].value
            list_data.append(datos) # agregar los datos a la lista de datos
    return list_data

def clean_balance_data(data):
    """
    La función es la encargada de limpiar los datos cuando estos no tienen el
    tipo de datos esperado, o tienen campo vacios que los hacen inválidos.

    Parámetros:
    -----------
    datos -> list - Lista de diccionarios con los datos a limpiar

    Retorna:
    --------

    """
    processed = list()
    for d in data:
        # Remover los diccionarios con GVO, GSV  NSV vacíos
        float_gov = isinstance(d['GOV'], float)
        float_gsv = isinstance(d['GSV'], float)
        float_nsv = isinstance(d['NSV'], float)
        if not float_gov and not float_gsv and not float_nsv:
            continue
        processed.append(d)
    return processed

def oil_sender_operation(data, operation_condition, operation_type):
    """
    Retorna un DataFrame con el total por tipo de crudo de determinado operación diario
    por cada remitente.

    Parámetros:
    -----------
    datos -> DataFrame - Contiene los datos del balance que serán usados para calcular
                        el todal de NSV recibido diario por cada empresa.

    Retorna:
    -------
    -> DataFrame - DataFrame que contiene el total diario de NSV por cada empresa,
                    una columna contiene los resultados de una empresa.
    """
    try:
        # Filtrar solo los datos cuya operación es un recibo
        recibidos = data[[operation_type in fila for fila in data['operacion']]]
        result = recibidos.groupby(['fecha', 'empresa'])[operation_condition].sum().unstack()
    except:
        result = pd.DataFrame()
    return result

def calculate_diff(data, operation_1, operation_2, operation_condition='NSV', company='GEOPARK'):
    """
    Retorna un DataFrame con el resultado de restar los datos de la caterogia_2 a la
    categoria_1 para cada día de operación por empresa y campo

    Parámetros:
    -----------
    datos  -> DataFrame - datos del balance
    categoria_1 -> str - cadena de caracteres con la primera categoría de donde vamos
                        a hacer la resta
    categoria_2 -> str - cadena de caracteres con al segunda categoría que le vamos
                        a restar a la primera categoría
    tipo_crudo -> str - Indica el tipo de crudo al cual se le van a sacar las diferentes
                        entre las dos categorías
    """
    # Filtrar los datos según las dos categorías a analizar
    filtro = (data['operacion'] ==  operation_1) | (data['operacion'] == operation_2)
    datos_filtrados = data[filtro]
    # Agrupar los datos para realizar las respectivas diferencias
    agrupados = datos_filtrados.groupby(['fecha', 'tipo crudo', 'empresa', 'operacion'])[operation_condition]
    # Separar los datos agrupados por operación
    op_1 = agrupados.sum().unstack().unstack().fillna(0)[operation_1]
    op_2 = agrupados.sum().unstack().unstack().fillna(0)[operation_2]
    # Calcular las diferencias
    diferencias = (op_1 - op_2).unstack().fillna(0)
    return diferencias[company]

def calculate_inventory_oil_type(data, company, operation_condition, initial_inventory_oil_type=None):
    """
    Retorna el inventario final por tipo crudo al sumar y restar las diferencias dadas
    en inventario_diario_campo del inventario_inicial_campo

    Parámetros:
    -----------
    Datos -> DataFrame - Contiene los datos de la operación diaria de las companies
    empresa -> str - Nombre de la empresa a la que le calculamos el inventario
    tipo_crudo -> str - Cadena de caracteres con el tipo de crudo a analizar
    inventario_inicial_campo -> dict - Diccionario con el inventario inicial para cada campo

    Retorna:
    --------
    pandas.core.series.Series -> Inventario por campo para la empresa indicada.
    """
    # calcular las diferencias para la empresa para determinado tipo de crudo
    diferencias = calculate_diff(data,
                                        'RECIBO POR REMITENTE TIGANA',
                                        'DESPACHO POR REMITENTE',
                                        operation_condition, company)
    # Agregar el inventario inicial por campo al dataframe de las diferencias
    if initial_inventory_oil_type is not None:
        diferencias = pd.concat([diferencias, pd.Series(initial_inventory_oil_type, dtype='float64')])
    # Eliminar las columnas que continen los datos de TIGANA y JACANA
    diferencias.drop(['JACANA ESTACION', 'TIGANA ESTACION'], inplace=True, axis=1)
    # Retornar el inventario final de los datos
    return np.round(diferencias.sum(), 2)

def calculate_total_inventory(inventory_oil_type):
    """
    Retorna la suma del inventario por campo.

    Parámetros:
    -----------
    pandas.core.series.Series - inventario por campo

    Retorna:
    float - Suma total de cada inventario por campo
    """
    return inventory_oil_type.sum()

def total_oil_detailed(datos, tipo_operacion):
    """
    Retorna un DataFrame con el total de Crudo según las condiciones de operación,
    el campo y el remitente.

    Parámetros:
    -----------
    datos -> DataFrame - Contiene los datos de crudo por tiempo de operación, campo y empresa.

    Retorna:
    --------
    total_crudo -> DataFrame - Totales por condiciones de operación, campo y remitente
    """
    # Filtrar los datos por tipo de operación
    total_crudo = datos[[tipo_operacion in fila for fila in datos['operacion']]]
    # Calcular los totales por condiciones de operación, campo y empresa
    total_crudo =  total_crudo.groupby(['empresa', 'tipo crudo'])[['GOV', 'GSV', 'NSV']].sum().unstack()
    # Se retorna el DataFrame con los totales, no sin antes remplazar los NaN por ceros.
    return total_crudo.fillna(0)

# Definición de funciones
def monthly_cumulated_oil_type(data, month, operation, company):
    """
    Retorna un DataFrame con el acumulado por campo para cada tipo de crudo
    en el mes, tipo de operación y empresa indicados
    """
    month_data = data[data['fecha'].dt.month == month]
    operation_data = month_data[month_data['operacion'] == operation]
    company_data = operation_data[operation_data['empresa'] == company]
    cumulated_month = company_data.groupby('tipo crudo')[['GOV', 'GSV', 'NSV']].sum()
    # Retornas los acumulados mensuales redondeados a 2 decimales
    return cumulated_month.round(2).reset_index()

def get_date_report(filename):
    return filename.split('Reportes')[-1].split()[2].split('.')[0]

def remove_entries_balance(filepath, filename):
    date_report = get_date_report(filename)
    df = pd.read_csv(filepath)
    # date_report = datetime.datetime.strftime(date_report)
    return df[df['fecha'] != date_report]
    

def estilo_celda(celda, background_color, font_color):
    """
    Agrega estilo a la celda indicada, y usa como background el color
    que recibe como parámetro
    """
    celda.fill = PatternFill('solid', fgColor=background_color)
    celda.alignment = Alignment(horizontal="center", vertical="center")
    thin = Side(border_style="thin", color="00000000")
    celda.border = Border(top=thin, left=thin, right=thin, bottom=thin)
    celda.font = Font(color=font_color, bold=True)

def agregar_estilos(hoja, filas, columna, background_color, font_color, header=True):
    """
    La función recibe una hoja de cálculo y se encarga de generar el estilo de la misma,
    agregando color a las celdas indicadas por las filas y la columna.
    """
    for fila in filas:
        celda = hoja.cell(row=fila, column=columna)
        # Dar estilo a la celda
        estilo_celda(celda, background_color, font_color)

        # Verificar si se trata de una cabecera
        if header:
            # Generar estilos para las celdas cuando se trata de una 
            # cabecera
            hoja.merge_cells(start_row=fila, start_column=columna, end_row=fila, end_column=columna + 3)
            # Estilo para las celdas restantes que no están en el merge
            for i in range(columna + 3, columna + 11):
                celda = hoja.cell(row=fila, column=i)
                estilo_celda(celda, background_color, font_color)
        else:
            # Agregar estilos para cuando se trata de celdas que no son parte
            # de la cabecera
            hoja.merge_cells(start_row=fila, start_column=columna, end_row=fila, end_column=columna + 10)

def write_data_monthly_report(data, month, year):
    """
    Escribir los datos acumulados por empresa y por tipo de operación en el
    ACTA para el mes indicado en un documento .xlsx
    """
    # Generar constante que almacena lo valores para la cabecera
    header = ['CAMPO','GOV (bls)','GSV (bls)','NSV (bls)','API @60ºF','S&W/Lab',
                '% Azufre','VISC 30 °C. /cSt']
    # Generar constante para cambiar los nombre de algunos valores
    names = {'RECIBO POR REMITENTE TIGANA': 'RECIBO POR REMITENTE EN TANQUE TK-780A',
                'PAREX': 'VERANO',
                'DESPACHO POR REMITENTE': 'DESPACHO POR REMITENTE',
                'ENTREGA POR REMITENTE': 'ENTREGA POR REMITENTE',
                'GEOPARK': 'GEOPARK',
                'RECIBO POR REMITENTE JACANA': 'RECIBO POR REMITENTE EN TANQUE 303'}
    # Declarar el objeto workbook
    book = Workbook()
    # Trabajar con la hoja activa.
    hoja = book.active
    hoja.insert_rows(1, amount=7)
    rows = 8
    filas_empresas = []
    filas_operaciones = []
    filas_cabecera = []
    try:
        for operation in data['operacion'].unique():
            hoja.append({6: names[operation]})
            rows += 1
            filas_operaciones.append(rows)
            for empresa in data['empresa'].unique():
                hoja.append({6: names[empresa]})
                rows += 1
                filas_empresas.append(rows)
                hoja.append({c + 6: value for c, value in enumerate(header)})
                rows += 1
                filas_cabecera.append(rows)
                acumulado = monthly_cumulated_oil_type(data, month, operation, empresa)
                acumulado['tipo crudo'] = [f'ACUMULADO MENSUAL {campo}' for campo in acumulado['tipo crudo']]
                for r in dataframe_to_rows(acumulado, index=False, header=False):
                    hoja.append({c + 6: value for c, value in enumerate(r)})
                    rows += 1

                if "DESPACHO" in operation:
                    hoja.append({6: ("ACUMULADO MENSUAL CRUDOS LLANOS 34 SEGMENTO I")})
                    hoja.append({6: ("ACUMULADO MENSUAL CRUDOS LLANOS 34 SEGMENTO II")})
                    hoja.append({6: ("ACUMULADO MENSUAL CRUDOS NO LLANOS 34 SEGMENTO I")})
                    hoja.append({6: ("ACUMULADO MENSUAL CRUDOS NO LLANOS 34 SEGMENTO I")})
                    rows+=4

                hoja.append(())
                rows += 1
            

        hoja.insert_rows(rows + 5)
        hoja.merge_cells(start_row=rows + 5, start_column=6, end_row=rows + 5, end_column=16)
        hoja["F" + str(rows + 5)] = "Gladys Maritza Fuentes Arciniegas"
        hoja["F" + str(rows + 5)].alignment = Alignment(horizontal="center", vertical="center")
        hoja["F" + str(rows + 5)].font = Font(bold=True)

        hoja.insert_rows(rows + 6)
        hoja.merge_cells(start_row=rows + 6, start_column=6, end_row=rows + 6, end_column=16)
        hoja["F" + str(rows + 6)].alignment = Alignment(horizontal="center", vertical="center")
        hoja["F" + str(rows + 6)] = "Coordinadora de Operaciones ODCA"

    except Exception as e:
        print(e)

    if not os.path.exists("../ReportesMensuales/Actas/"):
        os.mkdir("../ReportesMensuales/Actas/")

    hoja.insert_cols(7, amount=3)
    report_name = f'ACTA ODCA_{ months[ month - 1]}_{year}.xlsx'
    book.save(f"../ReportesMensuales/Actas/{ report_name }")
    return filas_cabecera, filas_empresas, filas_operaciones

def generate_report_ODCA(data, month, year):
    """
    Generar el acta con todos los datos requeridos y el estilo requerido
    """
    report_name = f'ACTA ODCA_{ months[ month - 1]}_{year}.xlsx'
    # Cargar los datos desde el balance y dar formato a las fechas
    data['fecha'] = pd.to_datetime(data['fecha'], format='%d-%m-%Y')
    # Escribir los datos en un documento .xlsx
    filas_cabecera, filas_empresas, filas_operaciones = write_data_monthly_report(data, month, year)
    # Cargar el documento generado anteriormente y seleccionar la hoja activa
    wb = load_workbook(f'../ReportesMensuales/Actas/{ report_name }')
    ws = wb.active
    # Agregar estilos al acta
    agregar_estilos(ws, filas_cabecera, 6, "000000", "FFFFFF")
    agregar_estilos(ws, filas_operaciones, 6, "FF0000", "FFFFFF", False)
    agregar_estilos(ws, filas_empresas, 6,"FFFFFF", "000000", False)

    # Add geopark logo
    img = Image("assets/logo_geopark.png")
    ws.add_image(img, 'A1')

    #Add title to the worksheet
    ws.merge_cells(start_row=2, start_column=6, end_row=6, end_column=16)
    ws["F2"] = """OLEODUCTO DEL CASANARE  (ODCA)
REPORTE DE OPERACIÓN MENSUAL"""
    ws["F2"].alignment = Alignment(horizontal="center", vertical="center")
    thin = Side(border_style="thin", color="00000000")
    ws["F2"].border = Border(top=thin, left=thin, right=thin, bottom=thin)
    ws["F2"].font = Font(bold=True)

    # Add date to worksheet
    ws.merge_cells(start_row=7, start_column=6, end_row=7, end_column=16)
    cell_date = ws['F7']
    cell_date.value = f"mes {months[ month - 1]}.{year}"
    cell_date.alignment = Alignment(horizontal="center", vertical="center")
    cell_date.font = Font(bold=True)

    # Cambiar el ancho de las columnas de los datos
    for i in range(10, 17):
        letter = get_column_letter(i)
        ws.column_dimensions[letter].width = 15

    if callback_context.triggered[0]['prop_id'] == "descargar-acta.n_clicks":
        wb.save(f'../ReportesMensuales/Actas/{ report_name }')
        return html.P(f'Se ha descargado el archivo: { report_name }')