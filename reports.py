from re import A
from tkinter import font
from turtle import back
from openpyxl import Workbook, load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import datetime
import pandas as pd

# Definición de funciones
def acumulado_mensual_campo(datos, mes, operacion, empresa):
    """
    Retorna un DataFrame con el acumulado por campo para cada tipo de crudo
    en el mes, tipo de operación y empresa indicados
    """
    datos_mes = datos[datos['fecha'].dt.month == mes]
    datos_operacion = datos_mes[datos_mes['operacion'] == operacion]
    datos_empresa = datos_operacion[datos_operacion['empresa'] == empresa]
    acumulado_mensual = datos_empresa.groupby('campo')[['GOV', 'GSV', 'NSV']].sum()
    # Retornas los acumulados mensuales redondeados a 2 decimales
    return acumulado_mensual.round(2).reset_index()

def estilo_celda(celda, background_color, font_color):
    """
    Agrega estilo a la celda indicada, y usa como background el color
    que recibe como parámetro
    """
    celda.fill = PatternFill('solid', fgColor=background_color)
    celda.alignment = Alignment(horizontal="center", vertical="center")
    thin = Side(border_style="thin", color="00000000")
    celda.border = Border(top=thin, left=thin, right=thin, bottom=thin)
    celda.font = Font(color=font_color, bold=True)

def agregar_estilos(hoja, filas, columna, background_color, font_color, header=True):
    """
    La función recibe una hoja de cálculo y se encarga de generar el estilo de la misma,
    agregando color a las celdas indicadas por las filas y la columna.
    """
    for fila in filas:
        celda = hoja.cell(row=fila, column=columna)
        # Dar estilo a la celda
        estilo_celda(celda, background_color, font_color)

        # Verificar si se trata de una cabecera
        if header:
            # Generar estilos para las celdas cuando se trata de una 
            # cabecera
            hoja.merge_cells(start_row=fila, start_column=columna, end_row=fila, end_column=columna + 3)
            # Estilo para las celdas restantes que no están en el merge
            for i in range(columna + 3, columna + 11):
                celda = hoja.cell(row=fila, column=i)
                estilo_celda(celda, background_color, font_color)
        else:
            # Agregar estilos para cuando se trata de celdas que no son parte
            # de la cabecera
            hoja.merge_cells(start_row=fila, start_column=columna, end_row=fila, end_column=columna + 10)

def escribir_datos(datos, mes):
    """
    Escribir los datos acumulados por empresa y por tipo de operación en el
    ACTA para el mes indicado en un documento .xlsx
    """
    # Generar constante que almacena lo valores para la cabecera
    cabecera = ['CAMPO','GOV (bls)','GSV (bls)','NSV (bls)','API @60ºF','S&W/Lab',
                '% Azufre','VISC 30 °C. /cSt']
    # Generar constante para cambiar los nombre de algunos valores
    nombres = {'RECIBO POR REMITENTE TIGANA': 'RECIBO POR REMITENTE EN TANQUE TK-780A',
                'PAREX': 'VERANO',
                'DESPACHO POR REMITENTE': 'DESPACHO POR REMITENTE',
                'ENTREGA POR REMITENTE': 'ENTREGA POR REMITENTE',
                'GEOPARK': 'GEOPARK'}
    # Declarar el objeto workbook
    libro = Workbook()
    # Trabajar con la hoja activa.
    hoja = libro.active
    hoja.insert_rows(1, amount=7)
    rows = 8
    filas_empresas = []
    filas_operaciones = []
    filas_cabecera = []
    for operacion in datos['operacion'].unique():
        hoja.append({6: nombres[operacion]})
        rows += 1
        filas_operaciones.append(rows)
        for empresa in datos['empresa'].unique():
            hoja.append({6: nombres[empresa]})
            rows += 1
            filas_empresas.append(rows)
            hoja.append({c + 6: value for c, value in enumerate(cabecera)})
            rows += 1
            filas_cabecera.append(rows)
            acumulado = acumulado_mensual_campo(datos, mes, operacion, 'GEOPARK')
            acumulado['campo'] = [f'ACUMULADO MENSUAL {campo}' for campo in acumulado['campo']]
            for r in dataframe_to_rows(acumulado, index=False, header=False):
                hoja.append({c + 6: value for c, value in enumerate(r)})
                rows += 1
            hoja.append(())
            rows += 1
    hoja.insert_cols(7, amount=3)
    libro.save("ACTA ODCA_" + str(mes) + '.xlsx')
    return filas_cabecera, filas_empresas, filas_operaciones

def generar_acta_ODCA(mes):
    """
    Generar el acta con todos los datos requeridos y el estilo requerido
    """
    # Cargar los datos desde el balance y dar formato a las fechas
    df = pd.read_csv('data/balance.csv')
    df['fecha'] = pd.to_datetime(df['fecha'], format='%d-%m-%Y')
    # Escribir los datos en un documento .xlsx
    filas_cabecera, filas_empresas, filas_operaciones = escribir_datos(df, mes)
    # Cargar el documento generado anteriormente y seleccionar la hoja activa
    wb = load_workbook('ACTA ODCA_' + str(mes) + '.xlsx')
    ws = wb.active
    # Agregar estilos al acta
    agregar_estilos(ws, filas_cabecera, 6, "000000", "FFFFFF")
    agregar_estilos(ws, filas_operaciones, 6, "FF0000", "FFFFFF", False)
    agregar_estilos(ws, filas_empresas, 6,"FFFFFF", "000000", False)
    # Cambiar el ancho de las columnas de los datos
    for i in range(10, 17):
        letter = get_column_letter(i)
        ws.column_dimensions[letter].width = 15
    wb.save('ACTA ODCA_' + str(mes) +'.xlsx')

if __name__ == "__main__":
    generar_acta_ODCA(2)
